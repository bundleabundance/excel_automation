import csv
import pandas as pd
import openpyxl
import os
from datetime import date

file_path = os.listdir()
file_name = "Telos Inventory Data 1-29-22.xlsx"

for file in file_path:
    if file.startswith("Business"):
        # columns 11, 15, 29, 33, 47, 49, 51 have formats
        # blue rgb level is (0, 0, 255) or 0000FF
        if file.endswith("gc-23-usa.csv"):
            wb = openpyxl.load_workbook(file_name)
            ws = wb["GC - Monthly Sale Input"]
            max_row = ws.max_row
            with open('BusinessReport-2-25-22-gc-23-usa.csv', 'r', encoding='utf-8', errors='ignore') as f:
                row_index = 3
                column_index = 1
                first_row = True
                for line in csv.reader(f):
                    if first_row:
                        first_row = False
                        continue
                    for cell in line:
                        if column_index > len(line):
                            column_index = 1
                            row_index += 1
                        ws.cell(row=max_row + row_index, column=column_index + 1, value=cell)
                        column_index += 1
                with open("BusinessReport-2-25-22-gc-23-can.csv", 'r', encoding="utf-8", errors='ignore') as f2:
                    row_index = 3
                    column_index = 1
                    first_row = True
                    for line in csv.reader(f2):
                        if first_row:
                            first_row = False
                            continue
                        for cell in line:
                            if column_index > len(line):
                                column_index = 1
                                row_index += 1
                            ws.cell(row=max_row + row_index, column=column_index + 19, value=cell)
                            column_index += 1
                wb.save(file_name)

