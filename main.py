import csv
import pandas as pd
import openpyxl
import os
from datetime import date


# get the date of today for handling the today's Excel file
# works for the first ten months because we exclude the first character (0) from the file name
date1 = date.today().strftime("%m-%d-%Y")[1:]
# print(date1)


# the function for iterating the files in the same directory as the running script
def iterate_files():

    file_path = os.listdir()
    file_name = "Telos Inventory Data 1-29-22.xlsx"
    # file_name == "Telos Inventory Data {}.xlsx".format(date1)

    # open the file to write
    # wb = openpyxl.load_workbook(file_name)

    # decide what to do according to the extensions of the files
    # all the operations are done according to the needs of the particular situation
    for file in file_path:

        if file == file_name or file == "main.py":
            # print("its working")
            continue

        elif file.endswith(".txt"):
            if file.startswith("GC - Fee Preview"):
                reader = pd.read_csv("GC - Fee Preview.txt", delimiter='\t', header=0, encoding='utf-8')
                reader.to_excel("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Fee Preview.xlsx", sheet_name='GC - Fee Preview', index=False, startrow=3)
                wb1 = openpyxl.load_workbook("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Fee Preview.xlsx")
                ws = wb1.active
                ws["A1"].value = "Reports > Fulfillment > Fee Preview (under Payments)"
                ws["A3"].value = date.today().strftime("%m-%d-%Y")
                wb1.save("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Fee Preview.xlsx")
            elif file.startswith("GC - AMZ Inventory"):
                reader = pd.read_csv("GC - AMZ Inventory.txt", delimiter='\t', header=0, encoding='utf-8')
                reader.to_excel("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - AMZ Inventory.xlsx", sheet_name='GC - AMZ Inventory', index=False, startrow=4, startcol=1)
                wb1 = openpyxl.load_workbook("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - AMZ Inventory.xlsx")
                ws = wb1.active
                ws["A1"].value = "Reports > Fulfillment > Manage FBA Inventory"
                ws["A2"].value = date.today().strftime("%m-%d-%Y")
                wb1.save("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - AMZ Inventory.xlsx")
            elif file.startswith("GC - Active Listings"):
                reader = pd.read_csv("GC - Active Listings.txt", delimiter='\t', header=0, encoding='utf-8')
                reader.to_excel("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Active Listings.xlsx", sheet_name='GC - Active Listings', index=False, startrow=3)
                wb1 = openpyxl.load_workbook("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Active Listings.xlsx")
                ws = wb1.active
                ws["A1"].value = "Inventory > Inventory Reports > Active Listings Report dropdown"
                ws["A3"].value = date.today().strftime("%m-%d-%Y")
                wb1.save("/Users/alper/Desktop/python/csv_excel/text_to_xl/GC - Active Listings.xlsx")

        elif file.endswith(".csv"):
            if file.startswith("Campaigns"):
                if file.endswith("gc.csv"):
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb["GC - Ad Export"]
                    with open('Campaigns_Feb_25_2022-gc.csv', 'r', encoding='utf-8', errors='ignore') as f:
                        row_index = 0
                        column_index = 1
                        for line in csv.reader(f):
                            if row_index == 0:
                                row_index += 1
                                continue
                            for cell in line:
                                if column_index > len(line):
                                    column_index = 1
                                    row_index += 1
                                ws.cell(row=row_index + 2, column=column_index + 1, value=cell)
                                column_index += 1
                        wb.save(file_name)

            if file.startswith("Business"):
                # columns 11, 15, 29, 33, 47, 49, 51 have formats
                # blue rgb level is (0, 0, 255) or 0000FF
                if file.endswith("gc-23-usa.csv"):
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb["GC - Monthly Sale Input"]
                    max_row = ws.max_row
                    with open('BusinessReport-2-25-22-gc-23-usa.csv', 'r', encoding='utf-8', errors='ignore') as f:
                        row_index = 3
                        column_index = 1
                        first_row = True
                        for line in csv.reader(f):
                            if first_row:
                                first_row = False
                                continue
                            for cell in line:
                                if column_index > len(line):
                                    column_index = 1
                                    row_index += 1
                                ws.cell(row=max_row + row_index, column=column_index + 1, value=cell)
                                column_index += 1
                        with open("BusinessReport-2-25-22-gc-23-can.csv", 'r', encoding="utf-8", errors='ignore') as f2:
                            row_index = 3
                            column_index = 1
                            first_row = True
                            for line in csv.reader(f2):
                                if first_row:
                                    first_row = False
                                    continue
                                for cell in line:
                                    if column_index > len(line):
                                        column_index = 1
                                        row_index += 1
                                    ws.cell(row=max_row + row_index, column=column_index + 19, value=cell)
                                    column_index += 1
                        wb.save(file_name)

        elif file.endswith(".xlsx"):
            if file.startswith("Sponsored Products"):
                wb1 = openpyxl.load_workbook("Sponsored Products Advertised product report-gc-8.xlsx")
                wb2 = openpyxl.load_workbook(file_name)
                wb2.iso_dates = True
                ws1 = wb1.active
                ws2 = wb2["GC - Sponsored Product"]
                max_row = ws2.max_row

                # for the indexing to not show
                cnt = -1
                for row in ws1:
                    cnt += 1
                    if cnt == 0:
                        continue
                    # for getting the next row to append
                    max_row += 1
                    # for formatting the date number of the first cells of the rows
                    first_cell = 0
                    for cell in row:
                        new_cell = ws2.cell(row=max_row, column=cell.col_idx, value=cell.value)
                        if first_cell == 0:
                            new_cell.number_format = "M.D.YYYY;@"

                            """ 
                            here is another example 
                            from openpyxl import load_workbook
                            from openpyxl.styles import NamedStyle
                            
                            xlsx_file = args.xlsx_file.name
                            
                            # opening:
                            wb = load_workbook(filename = xlsx_file)
                            
                            
                            # create date style:
                            date_style = NamedStyle(name='date_style', number_format='DD.MM.YYYY HH:MM:MM')
                            
                            # apply the style to the column H of the default sheet:
                            ws = wb.active
                            for row in ws[2:ws.max_row]:  # skip the header
                                cell = row[7]             # column H
                                cell.style = date_style
                            # saving:
                            wb.save(xlsx_file) 
                            """

                        first_cell += 1
                wb1.save("Sponsored Products Advertised product report-gc-8.xlsx")
                wb2.save(file_name)

    # save the file
    # wb.save(file_name)
    # wb.close()


iterate_files()

"""
import glob
from copy import copy
def merge_files():
    writer = pd.ExcelWriter("Output.xlsx")
    for file in glob.glob("*.xlsx"):
        (_, f_name) = os.path.split(file)
        (f_short_name, _) = os.path.splitext(f_name)
        df_excel = pd.read_excel(file, engine='openpyxl')
        df_excel.to_excel(writer, f_short_name, index=False)

    writer.save()
    
    writer = pd.ExcelWriter("output.xlsx")

    for filename in glob.glob("*.xlsx"):
        excel_file = pd.ExcelFile(filename)
        (_, f_name) = os.path.split(filename)
        (f_short_name, _) = os.path.splitext(f_name)
        for sheet_name in excel_file.sheet_names:
            df_excel = pd.read_excel(filename, sheet_name=sheet_name)
            df_excel.to_excel(writer, f_short_name, index=False)

    writer.save()
    """



"""
wb = openpyxl.load_workbook("Telos Inventory Data 1-29-22.xlsx")

# grab the active worksheet
ws = wb.active

with open('ff.csv', 'r', encoding='utf-8', errors='ignore') as f:
    for row in csv.reader(f):
        ws.append(row)
    # save the file
    wb.save('rr.xlsx')
"""

"""
with open('words.txt','r') as f:
    for line in f:
        for word in line.split():
           print(word)  
"""

"""
from openpyxl import load_workbook
wb = load_workbook('workbook.xlsx')
ws = wb[sheetname]
for idx, line in enumerate(csvfile):
    ws.cell(row=idx, column=0) = line[0]
    ws.cell(row=idx, column=1) = line[1]
wb.save("changed.xlsx")
"""
"""
sheet = wb["GC - Fee Preview"]
with open("GC - Fee Preview.txt", 'r') as f: # , encoding='utf-8' , errors='ignore'
    row_index = 0
    column_index = 1
    for line in f:
        for word in line.split():
            if column_index > len(line.split()):
                column_index = 1
                row_index += 1
            sheet.cell(row=row_index + 4, column=column_index).value = word
            column_index += 1
"""